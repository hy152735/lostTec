import openpyxl as px
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import sys

wb = px.load_workbook("template.xlsm")

# エテ名の定義
# key: エーテ界におけるエテの正式名称（エクセルに入力された名前)
# value: tenhoにおけるエテのハンドルネーム
ete_names = {
    '妖怪': '妖怪みむら',
    'マゾイ': 'マゾイ',
    '仏': 'こぼとけ',
    'hsmt': 'hsmt_ete',
    'UMD': 'tehutehu',
}

ws = wb['ヨンマ']

# エテハンドルネームとエクセルにおける列番号を対応付ける辞書
ete_columns = {}


#エクセルの1行目(名前行)をなめて、エテの名称と列番号を紐づける
# 行番号と列番号とを初期化
row = 1
column = 3

while True:
    # エクセルに入力された名前を取得
    ete_name = ws.cell(row=row, column=column).value
    if ete_name is None:
        break
    #エテのハンドルネームをkey、列番号をvalueとして定義
    ete_columns[ete_names[ete_name]] = column
    column += 1

log_file = sys.argv[1]
with open(log_file, "r", encoding="utf-8") as f:
    data = f.read()

def get_result(result_str):
    ete_name, result = result_str.split('(')
    ete_result = float(result.replace(')', ''))
    return ete_name, ete_result

# データの開始行は9行目から
row = 9
for line in data.split("\n"):
    ete_results = line.split(" ")
    one_game_result = {}
    # データの取得
    for i in range(6,10):
        result_str = ete_results[i]
        ete_name, ete_result = get_result(result_str)
        one_game_result[ete_name] = ete_result

        #五捨六入処理
        for key, value in one_game_result.items():
            #プラスの場合-0.1を、マイナスの場合そのまま四捨五入（roundは正確じゃないため、Decimal.quantizeを使用）
            pre_val = value - 0.1 if value >= 0 else value
            one_game_result[key] = int(Decimal(str(pre_val)).quantize(Decimal('0'), ROUND_HALF_UP))

        #合計が０になるよう、トップエテの値を補正
        top_ete = max(one_game_result, key=one_game_result.get)
        one_game_result[top_ete] -= sum(one_game_result.values())

    # NAGAが不存在の場合のみ、エクセルに追記
    for ete_name in one_game_result.keys():
        if "NAGA" in ete_name:
            break
    else:
        for ete_name, ete_result in one_game_result.items():
            column = ete_columns[ete_name]
            ws.cell(row=row, column=column).value = ete_result
        row += 1

dest_file = "%s.xlsm" % datetime.now().strftime('%Y-%m-%d')

wb.save(dest_file)
